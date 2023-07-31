import openpyxl
import warnings
from collections import OrderedDict


def get_secret(setting: str, secrets: dict):
    """Get secret setting or fail with ImproperlyConfigured"""
    try:
        return secrets[setting]
    except KeyError:
        raise ImproperlyConfigured("Set the {} setting".format(setting))


def suppressWarnings(f):
    warnings.simplefilter("ignore")
    return (f)


class upload_template:

    def __init__(
            self,
            path: str,
            lines: str = 'List of lines',
            instruction: str = 'Instruction',
            instruction_skip: int = 3
    ):
        @suppressWarnings
        def load_workbook(path: str):
            return openpyxl.load_workbook(path)

        # Load template to get info
        wb = load_workbook(path)

        # Get the column names from line sheet (lsheet)
        lsheet = wb[lines]
        # Get column names from the template
        for row in lsheet.iter_rows():
            self.field_labels = [cname.value for cname in row]
            break

        # Get default choices for each field from the instruction sheet
        # (isheet)
        isheet = wb[instruction]
        self.field_dict = OrderedDict()
        self.field_keys = list()
        
        # Get field names from the sheet (row #2)
        for nrow, row in enumerate(isheet.iter_rows()):
            if nrow == 1:
                for nkey, key in enumerate(row):
                    self.field_keys.append(key.value)
                    self.field_dict[key.value] = self.field_labels[nkey]

        choices = list()
        # Count the number of rows skipped before starting the rows containing
        # default choices
        skipped_rows = 0
        first_choice_row = True
        self.field_opts = OrderedDict()
        self.field_choices = OrderedDict()
        col_with_spec = []

        # Iterating over the rows and
        # getting value from each cell in row
        for row in isheet.iter_rows():
            # The first row is header;
            # the second is instruction
            if skipped_rows < instruction_skip:
                skipped_rows = skipped_rows + 1
                continue

            # Read default classes until no columns contain new classes
            # In the first choice row
            if first_choice_row:
                for ncell, cell in enumerate(row):
                    # If a column is non-empty
                    if cell.value is not None:
                        # Record the col number
                        col_with_spec.append(ncell)
                        val = str(cell.value)
                        
                        # Initiate the Ordered dictionary with
                        # the first choice
                        out_tuple = (val, val)
                        self.field_opts[self.field_keys[ncell]] = [out_tuple, ]
                        self.field_choices[self.field_keys[ncell]] = [val, ]

                first_choice_row = False                
                continue
            for ncol, cell in enumerate(row):
                # From the second choice row on, if the column had values in
                # it...
                if ncol in col_with_spec:
                    # If the row still have a value in it, add it to the 
                    # default choices
                    if cell.value is not None:
                        val = str(cell.value)
                        out_tuple = (val, val)
                        self.field_opts[self.field_keys[ncol]].append(out_tuple)
                        self.field_choices[self.field_keys[ncol]].append(val)
                        pass
                    # If there's no values, remove the column from examination
                    # in the next iteration
                    else:
                        col_with_spec.remove(ncol)

                # At the last column, examine if there's any column to check
                # on the next iteration. If not, stop the loop.
                if ncol == len(row) and len(col_with_spec) == 0:
                    break

    def get_field_index(self, query: str):
        if query in self.field_keys:
            return self.field_keys.index(query)
        else:
            error_msg = ' is not a valid field in the template.'
            raise ValueError(query + error_msg)
